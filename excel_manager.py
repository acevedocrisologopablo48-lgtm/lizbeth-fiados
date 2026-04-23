import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
import io

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

from config import EXCEL_PATH, SHEET_NAME, DATE_FORMAT


def _normalize(name: str) -> str:
    """Normaliza un nombre para comparación: minúsculas, sin tildes, sin espacios extra."""
    name = name.strip().lower()
    nfkd = unicodedata.normalize("NFKD", name)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _get_week_range(date: datetime) -> tuple[datetime, datetime]:
    """Retorna (lunes, domingo) de la semana que contiene la fecha dada."""
    monday = date - timedelta(days=date.weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday


def _ensure_workbook() -> openpyxl.Workbook:
    """Abre el Excel existente o crea uno nuevo con encabezados."""
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if SHEET_NAME not in wb.sheetnames:
            wb.create_sheet(SHEET_NAME)
            _write_headers(wb[SHEET_NAME])
        return wb

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _write_headers(ws)
    wb.save(EXCEL_PATH)
    return wb


def _write_headers(ws):
    """Escribe los encabezados con formato."""
    headers = ["Fecha", "Nombre", "Total Diario (S/)", "Total Semanal (S/)"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Anchos de columna
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22


def _find_row(ws, date_str: str, name: str) -> int | None:
    """Busca la fila donde coincide fecha y nombre. Retorna número de fila o None."""
    name_norm = _normalize(name)
    for row in range(2, ws.max_row + 1):
        cell_date = ws.cell(row=row, column=1).value
        cell_name = ws.cell(row=row, column=2).value
        if cell_date is None or cell_name is None:
            continue
        # La fecha puede estar como string o datetime
        if isinstance(cell_date, datetime):
            cell_date_str = cell_date.strftime(DATE_FORMAT)
        else:
            cell_date_str = str(cell_date)
        if cell_date_str == date_str and _normalize(str(cell_name)) == name_norm:
            return row
    return None


def _calc_weekly_total(ws, name: str, today: datetime) -> float:
    """Calcula el total semanal sumando todos los días de la semana para ese nombre."""
    monday, sunday = _get_week_range(today)
    name_norm = _normalize(name)
    total = 0.0
    for row in range(2, ws.max_row + 1):
        cell_date = ws.cell(row=row, column=1).value
        cell_name = ws.cell(row=row, column=2).value
        if cell_date is None or cell_name is None:
            continue
        if isinstance(cell_date, str):
            try:
                cell_date = datetime.strptime(cell_date, DATE_FORMAT)
            except ValueError:
                continue
        if _normalize(str(cell_name)) == name_norm and monday <= cell_date <= sunday:
            val = ws.cell(row=row, column=3).value
            if isinstance(val, (int, float)):
                total += val
    return total


def _update_weekly_totals(ws, name: str, today: datetime):
    """Actualiza la columna D (total semanal) para todas las filas del nombre en la semana."""
    monday, sunday = _get_week_range(today)
    name_norm = _normalize(name)
    weekly_total = _calc_weekly_total(ws, name, today)
    for row in range(2, ws.max_row + 1):
        cell_date = ws.cell(row=row, column=1).value
        cell_name = ws.cell(row=row, column=2).value
        if cell_date is None or cell_name is None:
            continue
        if isinstance(cell_date, str):
            try:
                cell_date = datetime.strptime(cell_date, DATE_FORMAT)
            except ValueError:
                continue
        if _normalize(str(cell_name)) == name_norm and monday <= cell_date <= sunday:
            ws.cell(row=row, column=4, value=round(weekly_total, 2))


def registrar_venta(nombre: str, monto: float) -> dict:
    """
    Registra una venta. Si el niño ya tiene entrada hoy, suma al monto.
    Retorna dict con info de la operación.
    """
    wb = _ensure_workbook()
    ws = wb[SHEET_NAME]
    today = datetime.now()
    date_str = today.strftime(DATE_FORMAT)
    nombre = nombre.strip().title()

    row = _find_row(ws, date_str, nombre)

    if row is not None:
        # UPDATE: sumar al monto existente
        monto_anterior = ws.cell(row=row, column=3).value or 0.0
        nuevo_total = round(monto_anterior + monto, 2)
        ws.cell(row=row, column=3, value=nuevo_total)
        _update_weekly_totals(ws, nombre, today)
        wb.save(EXCEL_PATH)
        return {
            "accion": "UPDATE",
            "nombre": nombre,
            "monto_anterior": monto_anterior,
            "monto_agregado": monto,
            "nuevo_total": nuevo_total,
        }
    else:
        # INSERT: nueva fila
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=today)
        ws.cell(row=new_row, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=new_row, column=2, value=nombre)
        ws.cell(row=new_row, column=3, value=round(monto, 2))
        ws.cell(row=new_row, column=3).number_format = "0.00"
        ws.cell(row=new_row, column=4).number_format = "0.00"
        _update_weekly_totals(ws, nombre, today)
        wb.save(EXCEL_PATH)
        return {
            "accion": "INSERT",
            "nombre": nombre,
            "monto": monto,
        }


def corregir_monto(nombre: str, nuevo_monto: float) -> dict:
    """Corrige (reemplaza) el monto diario de un niño para hoy."""
    wb = _ensure_workbook()
    ws = wb[SHEET_NAME]
    today = datetime.now()
    date_str = today.strftime(DATE_FORMAT)
    nombre = nombre.strip().title()

    row = _find_row(ws, date_str, nombre)
    if row is None:
        return {"error": f"No se encontró a '{nombre}' en el registro de hoy."}

    monto_anterior = ws.cell(row=row, column=3).value or 0.0
    ws.cell(row=row, column=3, value=round(nuevo_monto, 2))
    _update_weekly_totals(ws, nombre, today)
    wb.save(EXCEL_PATH)
    return {
        "accion": "CORREGIR",
        "nombre": nombre,
        "monto_anterior": monto_anterior,
        "nuevo_monto": nuevo_monto,
    }


def borrar_entrada(nombre: str) -> dict:
    """Elimina la entrada de un niño del día de hoy."""
    wb = _ensure_workbook()
    ws = wb[SHEET_NAME]
    today = datetime.now()
    date_str = today.strftime(DATE_FORMAT)
    nombre_title = nombre.strip().title()

    row = _find_row(ws, date_str, nombre_title)
    if row is None:
        return {"error": f"No se encontró a '{nombre_title}' en el registro de hoy."}

    monto = ws.cell(row=row, column=3).value or 0.0
    ws.delete_rows(row)
    _update_weekly_totals(ws, nombre_title, today)
    wb.save(EXCEL_PATH)
    return {
        "accion": "BORRAR",
        "nombre": nombre_title,
        "monto_eliminado": monto,
    }


def resumen_dia() -> list[dict]:
    """Retorna el resumen de todas las ventas del día de hoy."""
    wb = _ensure_workbook()
    ws = wb[SHEET_NAME]
    today = datetime.now()
    date_str = today.strftime(DATE_FORMAT)
    registros = []

    for row in range(2, ws.max_row + 1):
        cell_date = ws.cell(row=row, column=1).value
        if cell_date is None:
            continue
        if isinstance(cell_date, datetime):
            cell_date_str = cell_date.strftime(DATE_FORMAT)
        else:
            cell_date_str = str(cell_date)
        if cell_date_str == date_str:
            registros.append({
                "nombre": ws.cell(row=row, column=2).value,
                "total_diario": ws.cell(row=row, column=3).value or 0.0,
                "total_semanal": ws.cell(row=row, column=4).value or 0.0,
            })
    return registros


def resumen_semanal() -> list[dict]:
    """Retorna resumen semanal agrupado por nombre (semana actual)."""
    wb = _ensure_workbook()
    ws = wb[SHEET_NAME]
    today = datetime.now()
    monday, sunday = _get_week_range(today)
    acumulado: dict[str, float] = {}

    for row in range(2, ws.max_row + 1):
        cell_date = ws.cell(row=row, column=1).value
        cell_name = ws.cell(row=row, column=2).value
        if cell_date is None or cell_name is None:
            continue
        if isinstance(cell_date, str):
            try:
                cell_date = datetime.strptime(cell_date, DATE_FORMAT)
            except ValueError:
                continue
        if monday <= cell_date <= sunday:
            name = str(cell_name)
            monto = ws.cell(row=row, column=3).value or 0.0
            acumulado[name] = acumulado.get(name, 0.0) + monto

    return [{"nombre": k, "total_semanal": round(v, 2)} for k, v in sorted(acumulado.items())]


def editar_entrada(nombre_antiguo: str, nombre_nuevo: str, nuevo_monto: float) -> dict:
    """Edita nombre y/o monto de un registro de hoy."""
    wb = _ensure_workbook()
    ws = wb[SHEET_NAME]
    today = datetime.now()
    date_str = today.strftime(DATE_FORMAT)
    nombre_antiguo = nombre_antiguo.strip().title()
    nombre_nuevo = nombre_nuevo.strip().title()

    row = _find_row(ws, date_str, nombre_antiguo)
    if row is None:
        return {"error": f"No se encontró a '{nombre_antiguo}' en el registro de hoy."}

    monto_anterior = ws.cell(row=row, column=3).value or 0.0
    ws.cell(row=row, column=2, value=nombre_nuevo)
    ws.cell(row=row, column=3, value=round(nuevo_monto, 2))
    # Si el nombre cambió, recalcular totales del nombre viejo también
    if _normalize(nombre_nuevo) != _normalize(nombre_antiguo):
        _update_weekly_totals(ws, nombre_antiguo, today)
    _update_weekly_totals(ws, nombre_nuevo, today)
    wb.save(EXCEL_PATH)
    return {
        "accion": "EDITAR",
        "nombre_anterior": nombre_antiguo,
        "nombre_nuevo": nombre_nuevo,
        "monto_anterior": monto_anterior,
        "nuevo_monto": nuevo_monto,
    }


# ── Estilos para exportación ─────────────────────────────────────────────────

_COLOR_HEADER_BG   = "1F3864"   # azul oscuro — título principal
_COLOR_DATE_BG     = "2E75B6"   # azul medio  — cabecera de cada día
_COLOR_COL_HEADER  = "4472C4"   # azul claro  — fila de columnas
_COLOR_ROW_ODD     = "DEEAF1"   # azul muy pálido — fila impar
_COLOR_ROW_EVEN    = "FFFFFF"   # blanco — fila par
_COLOR_SUBTOTAL_BG = "BDD7EE"   # azul pastel — subtotal del día
_COLOR_TOTAL_BG    = "1F3864"   # azul oscuro — gran total
_COLOR_AMOUNT      = "1F5C99"   # texto de montos


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(bold=False, size=11, color="000000", name="Calibri") -> Font:
    return Font(name=name, bold=bold, size=size, color=color)


def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def generate_styled_export() -> io.BytesIO:
    """
    Genera un Excel estilizado con TODOS los datos acumulados,
    agrupados por fecha, con subtotales por día y gran total.
    Devuelve un BytesIO listo para enviar como archivo.
    """
    wb_src = _ensure_workbook()
    ws_src = wb_src[SHEET_NAME]

    # ── Leer y agrupar todos los datos por fecha ──────────────────────────
    from collections import defaultdict, OrderedDict
    days: dict[str, list[dict]] = defaultdict(list)

    for row in range(2, ws_src.max_row + 1):
        cell_date = ws_src.cell(row=row, column=1).value
        cell_name = ws_src.cell(row=row, column=2).value
        if cell_date is None or cell_name is None:
            continue
        if isinstance(cell_date, datetime):
            date_str = cell_date.strftime(DATE_FORMAT)
            date_obj = cell_date
        else:
            date_str = str(cell_date)
            try:
                date_obj = datetime.strptime(date_str, DATE_FORMAT)
            except ValueError:
                continue
        monto = ws_src.cell(row=row, column=3).value or 0.0
        days[date_str].append({
            "date_obj": date_obj,
            "nombre": str(cell_name),
            "monto": float(monto),
        })

    # Ordenar fechas cronológicamente
    sorted_days = OrderedDict(
        sorted(days.items(), key=lambda x: x[1][0]["date_obj"])
    )

    # ── Crear workbook de exportación ─────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro de Fiado"
    ws.sheet_view.showGridLines = False

    # Anchos de columna
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 22

    cur_row = 1

    # ── Fila de título principal ──────────────────────────────────────────
    ws.merge_cells(f"A{cur_row}:D{cur_row}")
    title_cell = ws.cell(row=cur_row, column=1,
                         value="🍽️  SISTEMA DE FIADO — LIZBETH")
    title_cell.font = _font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = _fill(_COLOR_HEADER_BG)
    title_cell.alignment = _align("center")
    title_cell.border = _border("medium")
    ws.row_dimensions[cur_row].height = 28
    cur_row += 1

    # Sub-título: fecha de generación
    ws.merge_cells(f"A{cur_row}:D{cur_row}")
    sub_cell = ws.cell(row=cur_row, column=1,
                       value=f"Generado el {datetime.now().strftime('%d/%m/%Y  %H:%M')}")
    sub_cell.font = _font(size=10, color="FFFFFF")
    sub_cell.fill = _fill(_COLOR_HEADER_BG)
    sub_cell.alignment = _align("center")
    sub_cell.border = _border()
    ws.row_dimensions[cur_row].height = 18
    cur_row += 1

    grand_total = 0.0

    for date_str, entries in sorted_days.items():
        date_obj = entries[0]["date_obj"]
        day_name_es = {
            0: "Lunes", 1: "Martes", 2: "Miércoles",
            3: "Jueves", 4: "Viernes", 5: "Sábado", 6: "Domingo",
        }[date_obj.weekday()]

        # ── Cabecera de fecha ─────────────────────────────────────────────
        cur_row += 1  # línea en blanco visual
        ws.row_dimensions[cur_row - 1].height = 6

        ws.merge_cells(f"A{cur_row}:D{cur_row}")
        date_header = ws.cell(row=cur_row, column=1,
                               value=f"📅  {day_name_es}  {date_str}")
        date_header.font = _font(bold=True, size=11, color="FFFFFF")
        date_header.fill = _fill(_COLOR_DATE_BG)
        date_header.alignment = _align("left")
        date_header.border = _border("medium")
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # ── Cabecera de columnas ──────────────────────────────────────────
        col_labels = ["#", "Nombre", "Monto del Día (S/)", "Total Semanal (S/)"]
        for c, lbl in enumerate(col_labels, 1):
            cell = ws.cell(row=cur_row, column=c, value=lbl)
            cell.font = _font(bold=True, size=10, color="FFFFFF")
            cell.fill = _fill(_COLOR_COL_HEADER)
            cell.alignment = _align("center")
            cell.border = _border()
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        # ── Filas de datos ────────────────────────────────────────────────
        day_total = 0.0
        for i, entry in enumerate(sorted(entries, key=lambda x: x["nombre"]), 1):
            bg = _COLOR_ROW_ODD if i % 2 == 1 else _COLOR_ROW_EVEN

            # Calcular total semanal para este nombre en esta semana
            monday, sunday = _get_week_range(date_obj)
            name_norm = _normalize(entry["nombre"])
            weekly = 0.0
            for r2 in range(2, ws_src.max_row + 1):
                cd = ws_src.cell(row=r2, column=1).value
                cn = ws_src.cell(row=r2, column=2).value
                if cd is None or cn is None:
                    continue
                if isinstance(cd, str):
                    try:
                        cd = datetime.strptime(cd, DATE_FORMAT)
                    except ValueError:
                        continue
                if _normalize(str(cn)) == name_norm and monday <= cd <= sunday:
                    val = ws_src.cell(row=r2, column=3).value or 0.0
                    weekly += float(val)

            cells_data = [i, entry["nombre"], entry["monto"], round(weekly, 2)]
            for c, val in enumerate(cells_data, 1):
                cell = ws.cell(row=cur_row, column=c, value=val)
                cell.fill = _fill(bg)
                cell.border = _border()
                cell.font = _font(size=10)
                if c == 1:
                    cell.alignment = _align("center")
                elif c == 2:
                    cell.alignment = _align("left")
                else:
                    cell.alignment = _align("right")
                    cell.number_format = '"S/ "#,##0.00'
                    cell.font = _font(size=10, color=_COLOR_AMOUNT, bold=(c == 3))

            day_total += entry["monto"]
            ws.row_dimensions[cur_row].height = 17
            cur_row += 1

        # ── Subtotal del día ──────────────────────────────────────────────
        ws.merge_cells(f"A{cur_row}:B{cur_row}")
        sub_label = ws.cell(row=cur_row, column=1, value="Subtotal del día")
        sub_label.font = _font(bold=True, size=10, color="1F3864")
        sub_label.fill = _fill(_COLOR_SUBTOTAL_BG)
        sub_label.alignment = _align("right")
        sub_label.border = _border()

        sub_val = ws.cell(row=cur_row, column=3, value=round(day_total, 2))
        sub_val.font = _font(bold=True, size=11, color="1F3864")
        sub_val.fill = _fill(_COLOR_SUBTOTAL_BG)
        sub_val.alignment = _align("right")
        sub_val.number_format = '"S/ "#,##0.00'
        sub_val.border = _border()

        empty_d = ws.cell(row=cur_row, column=4, value="")
        empty_d.fill = _fill(_COLOR_SUBTOTAL_BG)
        empty_d.border = _border()

        ws.row_dimensions[cur_row].height = 18
        grand_total += day_total
        cur_row += 1

    # ── Gran total ────────────────────────────────────────────────────────
    cur_row += 1
    ws.row_dimensions[cur_row - 1].height = 8

    ws.merge_cells(f"A{cur_row}:B{cur_row}")
    gt_label = ws.cell(row=cur_row, column=1, value="GRAN TOTAL ACUMULADO")
    gt_label.font = _font(bold=True, size=12, color="FFFFFF")
    gt_label.fill = _fill(_COLOR_TOTAL_BG)
    gt_label.alignment = _align("right")
    gt_label.border = _border("medium")

    gt_val = ws.cell(row=cur_row, column=3, value=round(grand_total, 2))
    gt_val.font = _font(bold=True, size=13, color="FFFFFF")
    gt_val.fill = _fill(_COLOR_TOTAL_BG)
    gt_val.alignment = _align("right")
    gt_val.number_format = '"S/ "#,##0.00'
    gt_val.border = _border("medium")

    gt_empty = ws.cell(row=cur_row, column=4, value="")
    gt_empty.fill = _fill(_COLOR_TOTAL_BG)
    gt_empty.border = _border("medium")

    ws.row_dimensions[cur_row].height = 24

    # Congelar las dos primeras filas
    ws.freeze_panes = "A3"

    # Guardar en buffer de memoria
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
